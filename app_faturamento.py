import streamlit as st
import pandas as pd
import openpyxl
import re
import json
import os
from io import BytesIO
import gspread
from google.oauth2.service_account import Credentials

# --- Configuração da Página ---
st.set_page_config(page_title="Portal de Faturamento", page_icon="📊", layout="wide")

# ==========================================================
# ECRÃ DE LOGIN (SENHA DA EQUIPE)
# ==========================================================
SENHA_CORRETA = "franquias2026"

if "logado" not in st.session_state:
    st.session_state.logado = False

if not st.session_state.logado:
    st.title("🔒 Acesso Restrito")
    senha_digitada = st.text_input("Digite a senha de acesso da equipe:", type="password")
    if st.button("Entrar"):
        if senha_digitada == SENHA_CORRETA:
            st.session_state.logado = True
            st.rerun()
        else:
            st.error("Senha incorreta!")
    st.stop()

# ==========================================================
# RESTO DO SISTEMA (SÓ RODA SE A SENHA ESTIVER CORRETA)
# ==========================================================

ARQUIVO_BACKUP = "backup_web.json"

# --- Configurações das Redes ---
REDES_CONFIG = {
    "La Brasa Burger": {
        "col_busca": "FRANQUIA",
        "categorias": [
            {"id": "iFood", "nome": "iFood", "is_sistema": False, "royalties": False},
            {"id": "Sistema", "nome": "Sistema / PDV", "is_sistema": True, "royalties": False},
            {"id": "Smaxi", "nome": "Smaxi", "is_sistema": False, "royalties": True},
            {"id": "Steak", "nome": "Steak", "is_sistema": False, "royalties": True},
            {"id": "F de Frango", "nome": "F de Frango", "is_sistema": False, "royalties": True}
        ],
        "usa_99_combo": True
    },
    "La Fruta Açaí": {
        "col_busca": "CIDADE",
        "categorias": [
            {"id": "La Fruta iFood", "nome": "La Fruta (iFood)", "is_sistema": False, "royalties": False},
            {"id": "La Fruta 99Food", "nome": "La Fruta (99Food)", "is_sistema": False, "royalties": False},
            {"id": "Lanches iFood", "nome": "Lanches e Sucos (iFood)", "is_sistema": False, "royalties": True},
            {"id": "Lanches 99Food", "nome": "Lanches e Sucos (99Food)", "is_sistema": False, "royalties": True},
            {"id": "La Fit iFood", "nome": "La Fit (iFood)", "is_sistema": False, "royalties": True},
            {"id": "La Fit 99Food", "nome": "La Fit (99Food)", "is_sistema": False, "royalties": True},
            {"id": "Baratim iFood", "nome": "Açaí Baratim (iFood)", "is_sistema": False, "royalties": True},
            {"id": "Baratim 99Food", "nome": "Açaí Baratim (99Food)", "is_sistema": False, "royalties": True},
            {"id": "Sistema", "nome": "Sistema / PDV", "is_sistema": True, "royalties": False}
        ],
        "usa_99_combo": False
    }
}

# --- Inicialização da Memória ---
if "dados_salvos" not in st.session_state: st.session_state.dados_salvos = {}
if "status_lojas" not in st.session_state: st.session_state.status_lojas = {}
if "lista_lojas" not in st.session_state: st.session_state.lista_lojas = []
if "template_bytes" not in st.session_state: st.session_state.template_bytes = None
if "rede_atual" not in st.session_state: st.session_state.rede_atual = "La Brasa Burger"

# --- Funções de Nuvem (Motor Oficial do Google) ---
def ligar_google_sheets():
    try:
        segredo = st.secrets["google_credentials"]
        
        if isinstance(segredo, str):
            cred_dict = json.loads(segredo)
        else:
            cred_dict = dict(segredo)
            
        if "private_key" in cred_dict:
            cred_dict["private_key"] = cred_dict["private_key"].replace("\\n", "\n")
            
        escopos = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        credenciais = Credentials.from_service_account_info(cred_dict, scopes=escopos)
        gc = gspread.authorize(credenciais)
        
        return gc.open("Base_Dados_Franquias")
    except Exception as e:
        st.error(f"⚠️ Erro detalhado ao conectar: {e}")
        return None

def enviar_para_nuvem(df, rede):
    sh = ligar_google_sheets()
    if not sh: return False, "Falha ao ligar ao Google Sheets."
    
    nome_aba = "Historico_La_Brasa" if rede == "La Brasa Burger" else "Historico_La_Fruta"
    try:
        ws = sh.worksheet(nome_aba)
    except:
        ws = sh.add_worksheet(title=nome_aba, rows="2000", cols="50")
    
    df = df.fillna("").astype(str)
    existentes = ws.get_all_values()
    
    if not existentes:
        ws.update(range_name='A1', values=[df.columns.tolist()] + df.values.tolist())
    else:
        cabecalho_atual = existentes[0]
        for col in df.columns:
            if col not in cabecalho_atual:
                cabecalho_atual.append(col)
        ws.update(range_name='A1', values=[cabecalho_atual])
        
        novas_linhas = []
        for _, row in df.iterrows():
            linha = []
            for col in cabecalho_atual:
                linha.append(row[col] if col in df.columns else "")
            novas_linhas.append(linha)
        ws.append_rows(novas_linhas)
        
    return True, "Enviado para a Nuvem com sucesso!"

def carregar_dados_nuvem(rede):
    sh = ligar_google_sheets()
    if not sh: 
        return pd.DataFrame()
        
    nome_aba = "Historico_La_Brasa" if rede == "La Brasa Burger" else "Historico_La_Fruta"
    try:
        ws = sh.worksheet(nome_aba)
        data = ws.get_all_records()
        return pd.DataFrame(data)
    except Exception as e:
        st.warning(f"A aba '{nome_aba}' ainda não contém dados gravados no Google Sheets.")
        return pd.DataFrame()

# --- Funções do Motor Matemático ---
def calcular_expressao(expr):
    if not expr: return 0.0
    expr = str(expr).replace(',', '.')
    expr_limpa = re.sub(r'[^\d.\+\-\*/]', '', expr)
    if not expr_limpa: return 0.0
    try: return float(eval(expr_limpa))
    except: return 0.0

def salvar_backup_local():
    backup = {
        "rede": st.session_state.rede_atual,
        "lista_lojas": st.session_state.lista_lojas,
        "status_lojas": st.session_state.status_lojas,
        "dados_salvos": st.session_state.dados_salvos
    }
    try:
        with open(ARQUIVO_BACKUP, "w", encoding="utf-8") as f:
            json.dump(backup, f, ensure_ascii=False, indent=4)
    except: pass

def carregar_backup_local():
    if os.path.exists(ARQUIVO_BACKUP):
        try:
            with open(ARQUIVO_BACKUP, "r", encoding="utf-8") as f:
                backup = json.load(f)
            st.session_state.rede_atual = backup.get("rede", "La Brasa Burger")
            st.session_state.lista_lojas = backup.get("lista_lojas", [])
            st.session_state.status_lojas = backup.get("status_lojas", {})
            st.session_state.dados_salvos = backup.get("dados_salvos", {})
            st.success("📂 Progresso recuperado com sucesso!")
        except Exception as e:
            st.error(f"Erro ao ler backup: {e}")
    else:
        st.warning("Nenhum arquivo de backup encontrado.")

def extrair_dados_com_openpyxl(file_bytes, rede):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    col_busca = REDES_CONFIG[rede]["col_busca"]
    dados_extraidos = {}
    
    de_para = {}
    if rede == "La Brasa Burger":
        de_para = {
            "Pedidos iFood": ["PEDIDOS IFOOD"], "Faturamento iFood": ["FATURAMENTO IFOOD"],
            "Pedidos Sistema": ["TOTAL PEDIDOS SISTEMA", "PEDIDOS SISTEMA"], "Faturamento Sistema": ["FATURAMENTO TOTAL SISTEMA", "FATURAMENTO SISTEMA"],
            "Pedidos Smaxi": ["PEDIDOS SMAXI BURGER", "PEDIDOS SMAXI"], "Faturamento Smaxi": ["FATURAMENTO TOTAL SMAXI BURGER", "FATURAMENTO SMAXI"],
            "Pedidos Steak": ["PEDIDOS STEAK BURGER", "PEDIDOS STEAK"], "Faturamento Steak": ["FATURAMENTO TOTAL STEAK BURGER", "FATURAMENTO STEAK"],
            "Pedidos F de Frango": ["PEDIDOS F DE FRANGO"], "Faturamento F de Frango": ["FATURAMENTO TOTAL F DE FRANGO", "FATURAMENTO F DE FRANGO"]
        }
    else:
        de_para = {
            "Faturamento La Fruta iFood": ["LA FRUTA IFOOD"], "Faturamento La Fruta 99Food": ["LA FRUTA 99FOOD"],
            "Faturamento Lanches iFood": ["LANCHES E SUCOS", "FATURAMENTO LANCHES"], "Faturamento Lanches 99Food": ["LANCHES & SUCOS 99FOOD"],
            "Faturamento La Fit iFood": ["LA FIT", "LA FIT IFOOD"], "Faturamento La Fit 99Food": ["LA FIT 99FOOD"],
            "Faturamento Baratim iFood": ["AÇAI BARATIM", "AÇAÍ BARATIM"], "Faturamento Baratim 99Food": ["AÇAI BARATIM 99FOOD", "AÇAÍ BARATIM 99FOOD"],
            "Pedidos Sistema": ["PEDIDOS SISTEMA", "TOTAL PEDIDOS SISTEMA"], "Faturamento Sistema": ["FATURAMENTO TOTAL SISTEMA", "FATURAMENTO SISTEMA"]
        }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        colunas_map = {}
        linha_cabecalho = -1
        nome_aba = sheet_name.upper()
        
        for row in range(1, 8):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str):
                    val_up = val.strip().upper()
                    if col_busca in val_up:
                        colunas_map["BUSCA"] = col
                        linha_cabecalho = row
                        
                    if rede == "La Brasa Burger" and "99FOOD" in val_up:
                        if "LA BRASA" in val_up: colunas_map["99_La Brasa"] = col
                        elif "SMAXI" in val_up: colunas_map["99_Smaxi"] = col
                        elif "STEAK" in val_up: colunas_map["99_Steak"] = col
                        elif "FRANGO" in val_up: colunas_map["99_F de Frango"] = col
                        else: colunas_map["99_Geral"] = col
                        
                    if rede == "La Fruta Açaí":
                        if val_up in ["IFOOD", "I FOOD", "I-FOOD"]:
                            if "FIT" in nome_aba: colunas_map["Faturamento La Fit iFood"] = col
                            elif "LANCHES" in nome_aba: colunas_map["Faturamento Lanches iFood"] = col
                            elif "BARATIM" in nome_aba: colunas_map["Faturamento Baratim iFood"] = col
                            elif "FRUTA" in nome_aba: colunas_map["Faturamento La Fruta iFood"] = col
                        elif val_up in ["99FOOD", "99 FOOD", "99-FOOD"]:
                            if "FIT" in nome_aba: colunas_map["Faturamento La Fit 99Food"] = col
                            elif "LANCHES" in nome_aba: colunas_map["Faturamento Lanches 99Food"] = col
                            elif "BARATIM" in nome_aba: colunas_map["Faturamento Baratim 99Food"] = col
                            elif "FRUTA" in nome_aba: colunas_map["Faturamento La Fruta 99Food"] = col
                        
                    for ch_interna, nomes in de_para.items():
                        if any(n in val_up for n in nomes): colunas_map[ch_interna] = col
            if "BUSCA" in colunas_map: break

        if "BUSCA" in colunas_map and linha_cabecalho != -1:
            idx_busca = colunas_map["BUSCA"]
            
            if rede == "La Fruta Açaí":
                novas_cols = {}
                for k_map, col_idx in colunas_map.items():
                    if k_map.startswith("Faturamento ") and "Sistema" not in k_map:
                        marca = k_map.replace("Faturamento ", "")
                        col_ped = col_idx - 1
                        if col_ped > 0:
                            cab = ws.cell(row=linha_cabecalho, column=col_ped).value
                            if isinstance(cab, str) and "PEDIDO" in cab.upper():
                                novas_cols[f"Pedidos {marca}"] = col_ped
                colunas_map.update(novas_cols)
            
            for row in range(linha_cabecalho + 1, ws.max_row + 1):
                celula = ws.cell(row=row, column=idx_busca).value
                if not celula: continue
                loja_excel = str(celula).strip()
                loja_lower = loja_excel.lower()
                
                if any(ignorado in loja_lower for ignorado in ["f de frango", "steak", "smaxi"]):
                    continue
                
                if loja_excel.replace("º", "").replace("°", "").strip().isdigit():
                    continue
                
                if loja_excel not in dados_extraidos:
                    dados_extraidos[loja_excel] = {"Franquia": loja_excel, "fechada": False}
                
                for k_map, col_idx in colunas_map.items():
                    if k_map != "BUSCA" and not k_map.startswith("99_"):
                        val_celula = ws.cell(row=row, column=col_idx).value
                        if isinstance(val_celula, (int, float)):
                            dados_extraidos[loja_excel][k_map] = val_celula

                if rede == "La Brasa Burger":
                    for m_key, m_name in [("99_La Brasa", "La Brasa Burger"), ("99_Smaxi", "Smaxi"), ("99_Steak", "Steak"), ("99_F de Frango", "F de Frango"), ("99_Geral", "Geral")]:
                        if m_key in colunas_map:
                            val_99 = ws.cell(row=row, column=colunas_map[m_key]).value
                            if val_99 and isinstance(val_99, (int, float)) and val_99 > 0:
                                dados_extraidos[loja_excel]["marca_99"] = m_name if m_name != "Geral" else "La Brasa Burger"
                                dados_extraidos[loja_excel]["faturamento_99"] = val_99
                                break
    return dados_extraidos

# --- BARRA LATERAL ---
st.sidebar.title("🛠️ Menu & Configurações")
rede_opcao = st.sidebar.selectbox("Selecione a Rede:", ["La Brasa Burger", "La Fruta Açaí"], index=0 if st.session_state.rede_atual == "La Brasa Burger" else 1)

mes_ref = st.sidebar.text_input("Mês/Ano de Referência (ex: 08/2026):", value="")

if rede_opcao != st.session_state.rede_atual:
    st.session_state.rede_atual = rede_opcao
    st.session_state.dados_salvos = {}
    st.session_state.status_lojas = {}
    st.session_state.lista_lojas = []
    st.session_state.template_bytes = None
    if "df_nuvem_cache" in st.session_state: del st.session_state["df_nuvem_cache"]
    st.rerun()

st.sidebar.markdown("---")
uploaded_file = st.sidebar.file_uploader("1. Carregar Planilha Molde (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    st.session_state.template_bytes = uploaded_file.getvalue()
    xls = pd.ExcelFile(uploaded_file)
    col_busca = REDES_CONFIG[st.session_state.rede_atual]["col_busca"]
    encontrou = False
    
    for aba in xls.sheet_names:
        for h in [2, 0, 1, 3, 4, 5]:
            df_temp = pd.read_excel(uploaded_file, sheet_name=aba, header=h)
            col_franquia = [c for c in df_temp.columns if col_busca in str(c).upper()]
            if col_franquia:
                lista_suja = df_temp[col_franquia[0]].dropna().astype(str).str.strip().unique().tolist()
                
                lojas_ignoradas = ["f de frango", "steak", "smaxi", "steak burger", "smaxi burger"]
                
                st.session_state.lista_lojas = sorted([
                    f for f in lista_suja 
                    if str(f).lower() != 'nan' and str(f).strip() != '' 
                    and str(f).lower() not in lojas_ignoradas
                    and not str(f).replace("º", "").replace("°", "").strip().isdigit()
                ])
                
                if not st.session_state.status_lojas:
                    st.session_state.status_lojas = {f: "Pendente" for f in st.session_state.lista_lojas}
                encontrou = True
                break
        if encontrou: break

    # BOTÃO DE IMPORTAÇÃO COM SINCRONIZAÇÃO INTELIGENTE
    if st.sidebar.button("📥 Importar Dados desta Planilha"):
        if not mes_ref:
            st.sidebar.error("⚠️ Preencha o 'Mês/Ano de Referência' antes de importar!")
        else:
            with st.spinner("Lendo o Excel e cruzando com a Nuvem..."):
                dados_lidos = extrair_dados_com_openpyxl(st.session_state.template_bytes, st.session_state.rede_atual)
                
                # Busca na nuvem o que já foi enviado
                if "df_nuvem_cache" not in st.session_state:
                    st.session_state.df_nuvem_cache = carregar_dados_nuvem(st.session_state.rede_atual)
                df_nuvem = st.session_state.df_nuvem_cache
                
                lojas_ja_feitas = []
                if not df_nuvem.empty:
                    col_mes = next((c for c in ["Mês Referência", "Mês/Ano", "Mês", "Período"] if c in df_nuvem.columns), None)
                    if col_mes:
                        ja_enviadas = df_nuvem[df_nuvem[col_mes].astype(str) == mes_ref]
                        if "Franquia" in ja_enviadas.columns:
                            lojas_ja_feitas = ja_enviadas["Franquia"].str.lower().tolist()

                # Marca as lojas inteligentes
                for loja, d in dados_lidos.items():
                    loja_match = next((l for l in st.session_state.lista_lojas if l.lower() == loja.lower()), None)
                    if loja_match:
                        st.session_state.dados_salvos[loja_match] = d
                        if loja_match.lower() in lojas_ja_feitas:
                            st.session_state.status_lojas[loja_match] = "Preenchida"
                        else:
                            st.session_state.status_lojas[loja_match] = "Pendente"
                            
                salvar_backup_local()
                st.sidebar.success(f"Pronto! Lista atualizada e sincronizada com a nuvem.")
                st.rerun()

st.sidebar.markdown("---")
if st.sidebar.button("📂 Restaurar Backup Local"):
    carregar_backup_local()
    st.rerun()

st.sidebar.markdown("---")
if st.sidebar.button("Sair (Bloquear Tela)"):
    st.session_state.logado = False
    st.rerun()

# --- ÁREA PRINCIPAL ---
st.title(f"📊 Gestão de Faturamento - {st.session_state.rede_atual}")

tab1, tab2 = st.tabs(["📝 Preenchimento de Lojas", "☁️ Base de Dados / Histórico na Nuvem"])

# ==============================================================================
# TAB 1: PREENCHIMENTO DE LOJAS
# ==============================================================================
with tab1:
    if not st.session_state.lista_lojas:
        st.info("👈 Faça o upload da **Planilha Molde (.xlsx)** no menu lateral para iniciar.")
    else:
        total = len(st.session_state.lista_lojas)
        preenchidas = sum(1 for s in st.session_state.status_lojas.values() if s == "Preenchida")
        fechadas = sum(1 for s in st.session_state.status_lojas.values() if s == "Fechada")
        pendentes = total - preenchidas - fechadas

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Preenchidas", f"{preenchidas} / {total}")
        m2.metric("Fechadas", fechadas)
        m3.metric("Pendentes", pendentes)
        st.progress(preenchidas / total if total > 0 else 0.0)
        st.markdown("---")

        opcoes_combo = []
        for loja in st.session_state.lista_lojas:
            st_loja = st.session_state.status_lojas.get(loja, "Pendente")
            if st_loja == "Preenchida": opcoes_combo.append(f"🟢 [✓] {loja}")
            elif st_loja == "Fechada": opcoes_combo.append(f"🔴 [X] {loja} (FECHADA)")
            else: opcoes_combo.append(f"⚪ [  ] {loja}")

        loja_raw = st.selectbox("Selecione a Unidade / Cidade:", ["-- Selecione --"] + opcoes_combo)

        if loja_raw != "-- Selecione --":
            nome_loja = loja_raw.replace("🟢 [✓] ", "").replace("🔴 [X] ", "").replace("⚪ [  ] ", "").replace(" (FECHADA)", "").strip()

            c_head1, c_head2 = st.columns([3, 1])
            with c_head1: st.subheader(f"Unidade: {nome_loja}")
            with c_head2:
                if st.button("🚫 Marcar Unidade como Fechada"):
                    st.session_state.status_lojas[nome_loja] = "Fechada"
                    st.session_state.dados_salvos[nome_loja] = {"fechada": True, "Mês Referência": mes_ref}
                    salvar_backup_local()
                    
                    df_fechada = pd.DataFrame([{"Mês Referência": mes_ref, "Franquia": nome_loja, "Status": "Fechada"}])
                    enviar_para_nuvem(df_fechada, st.session_state.rede_atual)
                    
                    st.rerun()

            dados_existentes = st.session_state.dados_salvos.get(nome_loja, {})

            with st.form(key=f"form_{nome_loja}"):
                categorias = REDES_CONFIG[st.session_state.rede_atual]["categorias"]
                inputs_coletados = {}

                for cat in categorias:
                    cid = cat["id"]
                    col_ped, col_fat, col_chk = st.columns([2, 3, 2])
                    val_ped = str(dados_existentes.get(f"Pedidos {cid}", "0"))
                    val_fat = str(dados_existentes.get(f"Faturamento {cid}", "0"))
                    val_int = dados_existentes.get(f"integra_{cid}", False)

                    with col_ped: ped_in = st.text_input(f"Pedidos - {cat['nome']}", value=val_ped, key=f"p_{cid}")
                    with col_fat: fat_in = st.text_input(f"Fat. (R$) - {cat['nome']}", value=val_fat, key=f"f_{cid}")
                    with col_chk:
                        st.write(""); st.write("")
                        int_in = False
                        if not cat["is_sistema"]: int_in = st.checkbox("Descontar do Sistema?", value=val_int, key=f"i_{cid}")
                    inputs_coletados[cid] = {"pedidos": ped_in, "fat": fat_in, "integra": int_in}

                inputs_99 = {}
                if REDES_CONFIG[st.session_state.rede_atual]["usa_99_combo"]:
                    st.markdown("**Opções 99Food (La Brasa)**")
                    c99_1, c99_2, c99_3, c99_4 = st.columns([2, 2, 3, 2])
                    with c99_1: m99 = st.selectbox("Marca 99Food", ["Nenhuma", "La Brasa Burger", "Smaxi", "Steak", "F de Frango"], key="m99_lb")
                    with c99_2: p99 = st.text_input("Pedidos 99", value=str(dados_existentes.get("pedidos_99", "0")), key="p99_lb")
                    with c99_3: f99 = st.text_input("Fat. 99 (R$)", value=str(dados_existentes.get("faturamento_99", "0")), key="f99_lb")
                    with c99_4:
                        st.write(""); st.write("")
                        i99 = st.checkbox("Descontar Sistema?", value=dados_existentes.get("integra_99", False), key="i99_lb")
                    inputs_99 = {"marca": m99, "pedidos": p99, "fat": f99, "integra": i99}

                if st.form_submit_button("💾 Salvar e Enviar p/ Nuvem", type="primary"):
                    dados_loja = {"Mês Referência": mes_ref, "Franquia": nome_loja, "Status": "Preenchida", "fechada": False}
                    id_sistema = [c["id"] for c in categorias if c["is_sistema"]][0]
                    ped_sis_bruto = int(calcular_expressao(inputs_coletados[id_sistema]["pedidos"]))
                    fat_sis_bruto = calcular_expressao(inputs_coletados[id_sistema]["fat"])
                    desc_fat = 0.0
                    desc_ped = 0
                    
                    for cat in categorias:
                        cid = cat["id"]
                        p = int(calcular_expressao(inputs_coletados[cid]["pedidos"]))
                        f = calcular_expressao(inputs_coletados[cid]["fat"])
                        dados_loja[f"Pedidos {cid}"] = p
                        dados_loja[f"Faturamento {cid}"] = f
                        dados_loja[f"integra_{cid}"] = inputs_coletados[cid]["integra"]
                        if not cat["is_sistema"] and inputs_coletados[cid]["integra"]:
                            desc_fat += f; desc_ped += p

                    if REDES_CONFIG[st.session_state.rede_atual]["usa_99_combo"]:
                        dados_loja["marca_99"] = inputs_99["marca"]
                        dados_loja["pedidos_99"] = int(calcular_expressao(inputs_99["pedidos"]))
                        dados_loja["faturamento_99"] = calcular_expressao(inputs_99["fat"])
                        dados_loja["integra_99"] = inputs_99["integra"]
                        if inputs_99["integra"]:
                            desc_fat += dados_loja["faturamento_99"]; desc_ped += dados_loja["pedidos_99"]

                    dados_loja["Pedidos Sistema"] = max(0, ped_sis_bruto - desc_ped)
                    dados_loja["Faturamento Sistema"] = max(0.0, fat_sis_bruto - desc_fat)

                    tot = sum(v for k, v in dados_loja.items() if "Faturamento" in k and isinstance(v, float) and "Bruto" not in k and "Sistema" not in k)
                    tot += dados_loja["Faturamento Sistema"]
                    dados_loja["Faturamento Total (R$)"] = tot

                    # --- MÁGICA INVISÍVEL: CÁLCULO DOS ROYALTIES DAS DARKS (4%) ---
                    total_royalties = 0.0
                    for cat in categorias:
                        if cat.get("royalties", False):
                            cid = cat["id"]
                            faturamento_marca = dados_loja.get(f"Faturamento {cid}", 0.0)
                            
                            if faturamento_marca > 0:
                                royalty = faturamento_marca * 0.04
                                dados_loja[f"Royalties {cat['nome']} (4%)"] = royalty
                                total_royalties += royalty
                                
                    if REDES_CONFIG[st.session_state.rede_atual]["usa_99_combo"]:
                        marca_99 = dados_loja.get("marca_99", "Nenhuma")
                        if marca_99 in ["Smaxi", "Steak", "F de Frango"]:
                            faturamento_99 = dados_loja.get("faturamento_99", 0.0)
                            if faturamento_99 > 0:
                                royalty_99 = faturamento_99 * 0.04
                                dados_loja[f"Royalties {marca_99} 99Food (4%)"] = royalty_99
                                total_royalties += royalty_99

                    if total_royalties > 0:
                        dados_loja["Total Royalties Darks (R$)"] = total_royalties

                    st.session_state.dados_salvos[nome_loja] = dados_loja
                    st.session_state.status_lojas[nome_loja] = "Preenchida"
                    salvar_backup_local()
                    
                    df_linha = pd.DataFrame([dados_loja])
                    sucesso, msg = enviar_para_nuvem(df_linha, st.session_state.rede_atual)
                    if sucesso: st.toast(f"✅ Salvo e enviado para o Google Sheets!")
                    else: st.error(msg)
                    
                    st.rerun()

        st.markdown("---")
        st.subheader("3. Exportação da Planilha Final")
        if st.button("🚀 Gerar e Baixar Excel Atualizado"):
            if not st.session_state.dados_salvos: st.warning("Nenhuma loja preenchida.")
            else:
                with st.spinner("Gerando arquivo Excel..."):
                    wb = openpyxl.load_workbook(BytesIO(st.session_state.template_bytes))
                    col_busca = REDES_CONFIG[st.session_state.rede_atual]["col_busca"]

                    de_para = {}
                    if st.session_state.rede_atual == "La Brasa Burger":
                        de_para = {
                            "Pedidos iFood": ["PEDIDOS IFOOD"], "Faturamento iFood": ["FATURAMENTO IFOOD"],
                            "Pedidos Sistema": ["TOTAL PEDIDOS SISTEMA", "PEDIDOS SISTEMA"], "Faturamento Sistema": ["FATURAMENTO TOTAL SISTEMA", "FATURAMENTO SISTEMA"],
                            "Pedidos Smaxi": ["PEDIDOS SMAXI BURGER", "PEDIDOS SMAXI"], "Faturamento Smaxi": ["FATURAMENTO TOTAL SMAXI BURGER", "FATURAMENTO SMAXI"],
                            "Pedidos Steak": ["PEDIDOS STEAK BURGER", "PEDIDOS STEAK"], "Faturamento Steak": ["FATURAMENTO TOTAL STEAK BURGER", "FATURAMENTO STEAK"],
                            "Pedidos F de Frango": ["PEDIDOS F DE FRANGO"], "Faturamento F de Frango": ["FATURAMENTO TOTAL F DE FRANGO", "FATURAMENTO F DE FRANGO"]
                        }
                    else:
                        de_para = {
                            "Faturamento La Fruta iFood": ["LA FRUTA IFOOD"], "Faturamento La Fruta 99Food": ["LA FRUTA 99FOOD"],
                            "Faturamento Lanches iFood": ["LANCHES E SUCOS", "FATURAMENTO LANCHES"], "Faturamento Lanches 99Food": ["LANCHES & SUCOS 99FOOD"],
                            "Faturamento La Fit iFood": ["LA FIT", "LA FIT IFOOD"], "Faturamento La Fit 99Food": ["LA FIT 99FOOD"],
                            "Faturamento Baratim iFood": ["AÇAI BARATIM", "AÇAÍ BARATIM"], "Faturamento Baratim 99Food": ["AÇAI BARATIM 99FOOD", "AÇAÍ BARATIM 99FOOD"],
                            "Pedidos Sistema": ["PEDIDOS SISTEMA", "TOTAL PEDIDOS SISTEMA"], "Faturamento Sistema": ["FATURAMENTO TOTAL SISTEMA", "FATURAMENTO SISTEMA"]
                        }

                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        colunas_map = {}
                        linha_cabecalho = -1
                        nome_aba = sheet_name.upper()
                        
                        for row in range(1, 7):
                            for col in range(1, ws.max_column + 1):
                                val = ws.cell(row=row, column=col).value
                                if isinstance(val, str):
                                    val_up = val.strip().upper()
                                    if col_busca in val_up:
                                        colunas_map["BUSCA"] = col
                                        linha_cabecalho = row
                                    for ch_interna, nomes in de_para.items():
                                        if any(n in val_up for n in nomes): colunas_map[ch_interna] = col
                            if "BUSCA" in colunas_map: break

                        if "BUSCA" in colunas_map and linha_cabecalho != -1:
                            idx_busca = colunas_map["BUSCA"]
                            
                            if st.session_state.rede_atual == "La Fruta Açaí":
                                novas_cols = {}
                                for k_map, col_idx in colunas_map.items():
                                    if k_map.startswith("Faturamento ") and "Sistema" not in k_map:
                                        marca = k_map.replace("Faturamento ", "")
                                        col_ped = col_idx - 1
                                        if col_ped > 0:
                                            cab = ws.cell(row=linha_cabecalho, column=col_ped).value
                                            if isinstance(cab, str) and "PEDIDO" in cab.upper():
                                                novas_cols[f"Pedidos {marca}"] = col_ped
                                colunas_map.update(novas_cols)

                            for row in range(linha_cabecalho + 1, ws.max_row + 1):
                                celula = ws.cell(row=row, column=idx_busca).value
                                if not celula: continue
                                loja_excel = str(celula).strip().lower()
                                
                                for fnome, dloja in st.session_state.dados_salvos.items():
                                    if fnome.lower() == loja_excel:
                                        if dloja.get("fechada", False):
                                            for cidx in colunas_map.values():
                                                if cidx != idx_busca: ws.cell(row=row, column=cidx).value = 0
                                        else:
                                            for k_map, col_idx in colunas_map.items():
                                                if k_map != "BUSCA" and k_map in dloja: ws.cell(row=row, column=col_idx).value = dloja[k_map]
                    output = BytesIO()
                    wb.save(output)
                    st.download_button("📥 Baixar Excel Atualizado", data=output.getvalue(), file_name=f"Faturamento_{st.session_state.rede_atual}.xlsx")

# ==============================================================================
# TAB 2: BASE DE DADOS / HISTÓRICO NA NUVEM (CARREGAMENTO AUTOMÁTICO)
# ==============================================================================
with tab2:
    st.header("☁️ Dados Consolidados do Google Sheets")
    st.write("Abaixo estão os dados reais gravados na nuvem. Os dados são carregados automaticamente assim que você entra na aba.")
    
    col_btn, _ = st.columns([1, 2])
    with col_btn:
        # BOTÃO CORRIGIDO COM "KEY" EXCLUSIVA
        if st.button("🔄 Forçar Recarregamento da Nuvem", key="btn_recarregar_nuvem"):
            if "df_nuvem_cache" in st.session_state:
                del st.session_state["df_nuvem_cache"]
            st.rerun()
        
    if "df_nuvem_cache" not in st.session_state:
        with st.spinner("Conectando e carregando dados da nuvem..."):
            st.session_state.df_nuvem_cache = carregar_dados_nuvem(st.session_state.rede_atual)
            
    df_nuvem = st.session_state.df_nuvem_cache
    
    if not df_nuvem.empty:
        col_mes = None
        for c in ["Mês Referência", "Mês/Ano", "Mês", "Período"]:
            if c in df_nuvem.columns:
                col_mes = c
                break
        
        if col_mes:
            meses_unicos = [str(m) for m in sorted(df_nuvem[col_mes].unique()) if str(m).strip() != ""]
            lista_opcoes_mes = ["Todos os Meses"] + meses_unicos
            mes_selecionado = st.selectbox("🗓️ Filtrar por Mês/Ano:", lista_opcoes_mes)
            if mes_selecionado != "Todos os Meses":
                df_exibir = df_nuvem[df_nuvem[col_mes].astype(str) == mes_selecionado].copy()
            else:
                df_exibir = df_nuvem.copy()
        else:
            df_exibir = df_nuvem.copy()
            
        # ======================================================================
        # NOVO MOTOR DE CÁLCULO DE CRESCIMENTO (%)
        # ======================================================================
        if col_mes and "Franquia" in df_exibir.columns:
            df_calc = df_nuvem.copy()
            df_calc['Data_Temp'] = pd.to_datetime(df_calc[col_mes], format='%m/%Y', errors='coerce')
            df_calc = df_calc.sort_values(by=['Franquia', 'Data_Temp'])
            
            # Remove duplicados no histórico para garantir o cálculo perfeito
            df_calc = df_calc.drop_duplicates(subset=['Franquia', col_mes])
            
            # Pega faturamento e pedidos, ignorando minúsculas
            colunas_analise = [c for c in df_calc.columns if ("faturamento" in c.lower() or "pedidos" in c.lower()) and "var." not in c.lower()]
            novas_colunas = []
            
            for col in colunas_analise:
                df_calc[col] = pd.to_numeric(df_calc[col], errors='coerce')
                # Puxa o dado do mês imediatamente anterior
                df_calc[f'{col}_Anterior'] = df_calc.groupby('Franquia')[col].shift(1)
                
                # Fórmula de crescimento
                divisor = df_calc[f'{col}_Anterior'].replace(0, pd.NA)
                var_col = f'Var. {col}'
                df_calc[var_col] = ((df_calc[col] - df_calc[f'{col}_Anterior']) / divisor) * 100
                
                def formatar_crescimento(val):
                    if pd.isna(val) or val == float('inf'): return "-"
                    if val > 0: return f"🟢 +{val:.1f}% 📈"
                    elif val < 0: return f"🔴 {val:.1f}% 📉"
                    else: return "⚪ 0.0%"
                    
                df_calc[var_col] = df_calc[var_col].apply(formatar_crescimento)
                novas_colunas.append(var_col)
                
            df_exibir = pd.merge(df_exibir, df_calc[['Franquia', col_mes] + novas_colunas], on=['Franquia', col_mes], how='left')

        df_visual = df_exibir.copy()
        
        # --- LIMPANDO COLUNAS "FEIAS" ---
        colunas_sujas = [c for c in df_visual.columns if "integra" in c.lower() or c.lower() in ["fechada", "status", "arquivo origem", "marca_99", "data_temp"]]
        df_visual = df_visual.drop(columns=colunas_sujas, errors='ignore')

        # --- ORDENADOR DE COLUNAS ---
        cols_order = []
        for c in df_visual.columns:
            if not str(c).startswith('Var. '):
                cols_order.append(c)
                var_col = f'Var. {c}'
                if var_col in df_visual.columns:
                    cols_order.append(var_col)
                    
        df_visual = df_visual[cols_order]

        # --- FORMATADOR DE NÚMEROS E MOEDAS ---
        for col in df_visual.columns:
            if ("faturamento" in col.lower() or "royalties" in col.lower()) and not str(col).startswith("Var."):
                df_visual[col] = pd.to_numeric(df_visual[col], errors='coerce').apply(
                    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notnull(x) else ""
                )
            elif "pedidos" in col.lower() and not str(col).startswith("Var."):
                df_visual[col] = pd.to_numeric(df_visual[col], errors='coerce').apply(
                    lambda x: f"{int(x)}" if pd.notnull(x) else ""
                )
            
        st.dataframe(df_visual, use_container_width=True)
        st.success(f"{len(df_exibir)} registros exibidos (de um total de {len(df_nuvem)} na nuvem).")
        
        csv_data = df_exibir.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Baixar Tabela Filtrada (CSV)", data=csv_data, file_name=f"faturamento_{st.session_state.rede_atual}_filtrado.csv", mime="text/csv")
    else:
        st.info("Ainda não existem registros gravados no Google Sheets para esta rede ou a ligação falhou.")

    st.markdown("---")
    st.header("📚 Construir Histórico de Meses Anteriores em Lote")
    arquivos_hist = st.file_uploader("Carregue as planilhas antigas", type=["xlsx"], accept_multiple_files=True)
    if st.button("☁️ Enviar Lote de Histórico para a Nuvem"):
        if not mes_ref:
            st.error("Por favor, preencha o campo 'Mês/Ano de Referência' na barra lateral antes de enviar.")
        elif arquivos_hist:
            linhas_hist = []
            with st.spinner("Lendo planilhas e gravando no Google Sheets..."):
                for arq in arquivos_hist:
                    dados_arq = extrair_dados_com_openpyxl(arq.getvalue(), st.session_state.rede_atual)
                    for loja, d in dados_arq.items():
                        linha = {"Mês Referência": mes_ref, "Arquivo Origem": arq.name, "Franquia": loja, "Status": "Histórico"}
                        tot_fat = 0.0
                        tot_ped = 0
                        for k, v in d.items():
                            if ("Faturamento" in k or "Pedidos" in k or "Royalties" in k) and isinstance(v, (int, float)):
                                linha[k] = v
                                if k.startswith("Faturamento ") and "Sistema" not in k and "Bruto" not in k: tot_fat += v
                                if k.startswith("Pedidos ") and "Sistema" not in k and "Bruto" not in k: tot_ped += v
                        
                        tot_fat += d.get("Faturamento Sistema", 0.0)
                        tot_ped += d.get("Pedidos Sistema", 0)
                        
                        # --- CORREÇÃO DO NOME DO TOTAL AQUI ---
                        linha["Faturamento Total (R$)"] = tot_fat
                        linha["Pedidos Total"] = tot_ped
                        
                        if tot_fat > 0 or tot_ped > 0: linhas_hist.append(linha)
            
            if linhas_hist:
                df_hist_total = pd.DataFrame(linhas_hist)
                sucesso, msg = enviar_para_nuvem(df_hist_total, st.session_state.rede_atual)
                if sucesso:
                    st.success("Histórico processado e enviado com sucesso para a nuvem!")
                    if "df_nuvem_cache" in st.session_state: del st.session_state["df_nuvem_cache"]
                    st.rerun()
                else:
                    st.error(msg)
        else:
            st.warning("Selecione pelo menos um arquivo Excel antigo.")